"""
Supabase 입력용 양식을 UTF-8 BOM CSV + 한글 안전 xlsx 로 생성합니다.

  py -3 -m pip install -r tools/import-templates/requirements-templates.txt
  py -3 tools/import-templates/build_templates.py

- 모든 CSV: UTF-8 BOM (엑셀에서 열 때 한글 깨짐 방지)
- quiz_import_template.xlsx: 동일 내용을 시트별로 넣음(바이너리 xlsx는 UTF-8 유니코드 문자열을 그대로 보관)
"""
from __future__ import annotations

import csv
import io
import json
import subprocess
import sys
from pathlib import Path

DIR = Path(__file__).resolve().parent
ROOT = DIR.parents[1]  # ox-quiz-app/


def write_utf8_sig_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)


def run_question_gen() -> None:
    script = DIR / "_gen_quiz_questions_template.py"
    r = subprocess.run([sys.executable, str(script)], cwd=str(ROOT))
    if r.returncode != 0:
        raise SystemExit("문항 템플릿 생성 실패 (_gen_quiz_questions_template.py)")


def write_static_templates() -> None:
    write_utf8_sig_csv(
        DIR / "quiz_subjects.template.csv",
        [
            ["id", "name", "sort_order", "is_active"],
            ["", "해양경찰학개론", "0", "true"],
        ],
    )
    write_utf8_sig_csv(
        DIR / "quiz_units.template.csv",
        [
            ["id", "subject_id", "parent_unit_id", "name", "sort_order", "is_active"],
            ["", "<SUBJECT_UUID>", "", "해양경찰의 개념", "0", "true"],
            ["", "<SUBJECT_UUID>", "<MAJOR_UNIT_UUID>", "해양경찰의 역사", "0", "true"],
        ],
    )
    write_utf8_sig_csv(
        DIR / "quiz_flat_entry.template.csv",
        [
            [
                "subject_name",
                "subject_sort_order",
                "major_unit_name",
                "major_sort_order",
                "minor_unit_name",
                "minor_sort_order",
                "pack_no",
                "question_sort",
                "question",
                "choice_text",
                "body",
                "answer",
                "explanation",
                "is_active",
            ],
            [
                "해양경찰학개론",
                "0",
                "해양경찰의 개념",
                "0",
                "해양경찰의 역사",
                "0",
                "301",
                "0",
                "다음은 해양경찰의 변천사를 설명한 것이다. ( ) 안에 들어갈 말은?",
                "㉠ 1953 ㉡ 1996",
                "문제: …\n\n선지: …",
                "TRUE",
                "정답 ①.",
                "true",
            ],
        ],
    )


def write_usage_txt_bom() -> None:
    path = DIR / "00_컬럼설명.txt"
    text = path.read_text(encoding="utf-8-sig")
    path.write_text("\ufeff" + text.lstrip("\ufeff"), encoding="utf-8")


def build_xlsx() -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("openpyxl 미설치 — xlsx 는 건너뜁니다. 설치: py -3 -m pip install -r tools/import-templates/requirements-templates.txt")
        return None

    json_path = DIR / "quiz_questions.rows.json"
    qdata = json.loads(json_path.read_text(encoding="utf-8"))
    headers: list[str] = qdata["headers"]
    qrows: list[dict] = qdata["rows"]

    wb = Workbook()
    # 사용법
    ws0 = wb.active
    ws0.title = "사용법"
    usage = [
        ["OX 퀴즈 Supabase 입력 양식"],
        [],
        ["한글 처리"],
        ["· CSV 파일은 모두 UTF-8 BOM 으로 저장되어 있습니다. 메모장/엑셀에서 열 때 깨지면「데이터 → 텍스트/CSV」에서 원본 UTF-8 을 지정하세요."],
        ["· xlsx 는 유니코드(UTF-16 내부)로 한글을 보관합니다. 다른 PC로 옮길 때도 .xlsx 로 주고받는 것이 가장 안전합니다."],
        [],
        ["시트 구성"],
        ["· quiz_subjects / quiz_units / quiz_questions / quiz_flat_entry = DB 또는 평면 입력 예시"],
        ["· quiz_questions 예시는 data.js 의 MDP_HISTORY_19G01_STEM 과 동기화됩니다."],
        [],
        ["unit_id"],
        ["· 예시 UUID 는 supabase/seed_haeyang_mdp.sql 의「해양경찰의 역사」소단원과 동일합니다. 본인 DB에 맞게 바꾸세요."],
    ]
    for i, row in enumerate(usage, start=1):
        for j, val in enumerate(row, start=1):
            c = ws0.cell(row=i, column=j, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 100

    def sheet_from_matrix(name: str, matrix: list[list[str]], header_bold: bool = True) -> None:
        ws = wb.create_sheet(title=name[:31])
        bold = Font(bold=True)
        for r, row in enumerate(matrix, start=1):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if r == 1 and header_bold:
                    cell.font = bold
        ws.freeze_panes = "A2"

    # subjects / units / flat from disk (방금 쓴 UTF-8 내용)
    for fname, title in [
        ("quiz_subjects.template.csv", "quiz_subjects"),
        ("quiz_units.template.csv", "quiz_units"),
        ("quiz_flat_entry.template.csv", "quiz_flat_entry"),
    ]:
        p = DIR / fname
        raw = p.read_text(encoding="utf-8-sig")
        rows = list(csv.reader(io.StringIO(raw)))
        sheet_from_matrix(title, rows)

    from openpyxl.utils import get_column_letter

    wsq = wb.create_sheet(title="quiz_questions")
    for c, h in enumerate(headers, start=1):
        cell = wsq.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, rowdict in enumerate(qrows, start=2):
        for ci, h in enumerate(headers, start=1):
            v = rowdict.get(h, "")
            wsq.cell(row=ri, column=ci, value=v).alignment = Alignment(wrap_text=True, vertical="top")
    wsq.freeze_panes = "A2"
    widths = {"id": 38, "unit_id": 38, "question": 50, "choice_text": 22, "body": 60, "explanation": 40}
    for ci, h in enumerate(headers, start=1):
        letter = get_column_letter(ci)
        wsq.column_dimensions[letter].width = min(widths.get(h, 16), 60)

    out_xlsx = DIR / "quiz_import_template.xlsx"
    wb.save(out_xlsx)
    print("wrote", out_xlsx)
    return out_xlsx


def main() -> None:
    write_static_templates()
    write_usage_txt_bom()
    run_question_gen()
    build_xlsx()
    print("완료: UTF-8 BOM CSV + (openpyxl 있으면) quiz_import_template.xlsx")


if __name__ == "__main__":
    main()
